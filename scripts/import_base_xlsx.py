import os
import sys
import json
import re
import argparse
from datetime import datetime
from decimal import Decimal, InvalidOperation
import openpyxl

try:
    import psycopg
except ImportError:
    psycopg = None

EXPECTED_UNIT_COUNT = 163
MIN_SAFE_COUNT = 150
MAX_SAFE_COUNT = 200
MONEY_QUANT = Decimal("0.01")

def normalize_digits(value):
    if value is None:
        return ""
    if isinstance(value, Decimal) and value == value.to_integral_value():
        value = int(value)
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    decimal_string = re.fullmatch(r'(\d+)\.0+', text)
    if decimal_string:
        text = decimal_string.group(1)
    return re.sub(r'\D', '', text)

def normalize_cnpj(cnpj_val):
    digits = normalize_digits(cnpj_val)
    if not digits: return None
    if len(digits) == 13:
        return digits.zfill(14)
    if len(digits) == 14:
        return digits
    raise ValueError(f"CNPJ inválido detectado (tamanho: {len(digits)}): {cnpj_val}")

def normalize_inep(inep_val):
    digits = normalize_digits(inep_val)
    if not digits: return None
    if len(digits) == 8:
        return digits
    raise ValueError(f"INEP inválido detectado (tamanho: {len(digits)}): {inep_val}")

def parse_money_ptbr(value):
    if value is None or str(value).strip() == "":
        return Decimal("0.00")

    if isinstance(value, (int, float, Decimal)):
        amount = Decimal(str(value))
        if amount < 0:
            raise ValueError(f"Valor financeiro negativo: {value}")
        return amount

    val_str = str(value).strip()
    val_str = val_str.replace("R$", "").strip()
    val_str = re.sub(r'\s+', '', val_str)

    if "," in val_str and "." in val_str:
        val_str = val_str.replace(".", "")
        val_str = val_str.replace(",", ".")
    elif "," in val_str:
        val_str = val_str.replace(",", ".")
    elif re.fullmatch(r'\d{1,3}(\.\d{3})+', val_str):
        val_str = val_str.replace(".", "")
        
    try:
        amount = Decimal(val_str)
    except InvalidOperation:
        raise ValueError(f"Valor financeiro inválido: {value}")
    if amount < 0:
        raise ValueError(f"Valor financeiro negativo: {value}")
    return amount

def decimal_to_str(d):
    return str(d.quantize(MONEY_QUANT))

def get_col_index(header_map, possible_names):
    for name in possible_names:
        if name in header_map:
            return header_map[name]
    return None

def process_file(filepath, exercicio=2026, programa='basico', apply=False, approve_divergent_count=False):
    print(f"Lendo {filepath} para exercício {exercicio} / {programa} (Apply: {apply})...")
    
    if not os.path.exists(filepath):
        print(f"ERRO: Arquivo {filepath} não encontrado.")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if 'BASE' not in wb.sheetnames:
        print("ERRO: Aba 'BASE' não encontrada na planilha.")
        sys.exit(1)
        
    ws = wb['BASE']
    
    header = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            header[str(cell.value).strip().upper()] = idx
            
    col_designacao = get_col_index(header, ['DESIGNAÇÃO', 'DESIGNACAO'])
    col_nome = get_col_index(header, ['NOME'])
    col_cnpj = get_col_index(header, ['CNPJ'])
    col_inep = get_col_index(header, ['INEP'])
    col_diretor = get_col_index(header, ['DIRETOR'])
    col_endereco = get_col_index(header, ['ENDEREÇO', 'ENDERECO'])
    col_agencia = get_col_index(header, ['AGÊNCIA', 'AGENCIA'])
    col_conta = get_col_index(header, ['CONTA CORRENTE', 'CONTA'])
    col_rep_cust = get_col_index(header, ['REPROGRAMADO CUSTEIO'])
    col_rep_cap = get_col_index(header, ['REPROGRAMADO CAPITAL'])
    col_p1_cust = get_col_index(header, ['1 PARCELA CUSTEIO', '1ª PARCELA CUSTEIO', '1A PARCELA CUSTEIO'])
    col_p1_cap = get_col_index(header, ['1 PARCELA CAPITAL', '1ª PARCELA CAPITAL', '1A PARCELA CAPITAL'])
    col_p2_cust = get_col_index(header, ['2 PARCELA CUSTEIO', '2ª PARCELA CUSTEIO', '2A PARCELA CUSTEIO'])
    col_p2_cap = get_col_index(header, ['2 PARCELA CAPITAL', '2ª PARCELA CAPITAL', '2A PARCELA CAPITAL'])

    if col_designacao is None or col_nome is None:
        print("ERRO: Colunas obrigatórias (Designação/Nome) não encontradas.")
        sys.exit(1)

    financial_columns = {
        'REPROGRAMADO CUSTEIO': col_rep_cust,
        'REPROGRAMADO CAPITAL': col_rep_cap,
        '1 PARCELA CUSTEIO': col_p1_cust,
        '1 PARCELA CAPITAL': col_p1_cap,
        '2 PARCELA CUSTEIO': col_p2_cust,
        '2 PARCELA CAPITAL': col_p2_cap,
    }
    missing_financial = [name for name, col in financial_columns.items() if col is None]
    if missing_financial:
        print(f"ERRO: Colunas financeiras obrigatórias ausentes: {', '.join(missing_financial)}")
        sys.exit(1)

    parsed_data = []
    skipped = 0
    errors_list = []
    seen_designacoes = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            designacao = str(row[col_designacao]).strip() if row[col_designacao] else None
            nome = str(row[col_nome]).strip() if row[col_nome] else None
            
            if not designacao or not nome or designacao == 'None':
                skipped += 1
                continue

            if '—' in designacao or ' - ' in designacao:
                raise ValueError("Designação contaminada com rótulo/nome da escola")
            if designacao in seen_designacoes:
                raise ValueError(f"Designação duplicada; primeira ocorrência na linha {seen_designacoes[designacao]}")
            seen_designacoes[designacao] = row_idx
                
            cnpj_val = normalize_cnpj(row[col_cnpj] if col_cnpj is not None else None)
            inep_val = normalize_inep(row[col_inep] if col_inep is not None else None)
            diretor_val = str(row[col_diretor]).strip() if col_diretor is not None and row[col_diretor] else None
            endereco_val = str(row[col_endereco]).strip() if col_endereco is not None and row[col_endereco] else None
            agencia_val = str(row[col_agencia]).strip() if col_agencia is not None and row[col_agencia] else None
            conta_val = str(row[col_conta]).strip() if col_conta is not None and row[col_conta] else None

            # Construir source_payload detalhado para auditoria
            source_payload = {
                '__row_number': row_idx,
                **{k: str(row[v]) if row[v] is not None else None for k, v in header.items()}
            }

            payload = {
                'unidade': {
                    'designacao': designacao,
                    'nome': nome,
                    'inep': inep_val,
                    'cnpj': cnpj_val,
                    'diretor': diretor_val,
                    'endereco': endereco_val,
                    'agencia': agencia_val,
                    'conta_corrente': conta_val,
                    'source_payload': source_payload
                },
                'financeiro': {
                    'exercicio': exercicio,
                    'programa': programa,
                    'reprogramado_custeio': decimal_to_str(parse_money_ptbr(row[col_rep_cust] if col_rep_cust is not None else 0)),
                    'reprogramado_capital': decimal_to_str(parse_money_ptbr(row[col_rep_cap] if col_rep_cap is not None else 0)),
                    'parcela_1_custeio': decimal_to_str(parse_money_ptbr(row[col_p1_cust] if col_p1_cust is not None else 0)),
                    'parcela_1_capital': decimal_to_str(parse_money_ptbr(row[col_p1_cap] if col_p1_cap is not None else 0)),
                    'parcela_2_custeio': decimal_to_str(parse_money_ptbr(row[col_p2_cust] if col_p2_cust is not None else 0)),
                    'parcela_2_capital': decimal_to_str(parse_money_ptbr(row[col_p2_cap] if col_p2_cap is not None else 0))
                }
            }
            parsed_data.append(payload)
        except Exception as e:
            err_msg = f"Linha {row_idx}: {str(e)}"
            print(f"Erro na validação - {err_msg}")
            errors_list.append({'linha': row_idx, 'erro': str(e)})

    total_valid = len(parsed_data)
    if apply:
        if errors_list:
            print(f"ERRO DE VALIDAÇÃO: {len(errors_list)} linha(s) inválida(s). Corrija a BASE antes de usar --apply.")
            sys.exit(1)
        if not (MIN_SAFE_COUNT <= total_valid <= MAX_SAFE_COUNT):
            print(f"ERRO DE SEGURANÇA: Contagem de unidades válidas ({total_valid}) fora do limite seguro ({MIN_SAFE_COUNT}-{MAX_SAFE_COUNT}).")
            sys.exit(1)
        if total_valid != EXPECTED_UNIT_COUNT and not approve_divergent_count:
            print(
                f"ERRO DE APROVAÇÃO: Contagem de unidades válidas ({total_valid}) diferente do esperado "
                f"({EXPECTED_UNIT_COUNT}). Reexecute com --approve-divergent-count somente após aprovação humana."
            )
            sys.exit(1)

    if not apply:
        os.makedirs('data/output', exist_ok=True)
        with open('data/output/import_preview.json', 'w', encoding='utf-8') as f:
            json.dump(parsed_data, f, ensure_ascii=False, indent=2)
            
        report = f"""# Relatório de Importação (Modo DRY-RUN)
Data: {datetime.now().isoformat()}
Arquivo: {filepath}
Exercício: {exercicio} / {programa}
Válidas: {total_valid}
Puladas: {skipped}
Erros de Validação: {len(errors_list)}
Modo: PREVIEW ONLY (Nenhuma alteração no banco)"""
        with open('data/output/import_report.md', 'w', encoding='utf-8') as f:
            f.write(report)
            if errors_list:
                f.write("\n\n## Erros de Validação\n")
                f.write(json.dumps(errors_list, ensure_ascii=False, indent=2))
        print(f"Dry-run concluído. {total_valid} linhas prontas. Relatório salvo em data/output/.")
        sys.exit(0)

    # Modo Apply: usa DATABASE_URL para gravar em transação única.
    db_url = os.environ.get('DATABASE_URL')

    if not db_url:
        print("ERRO: Variável DATABASE_URL não encontrada para o modo --apply.")
        sys.exit(1)
    if psycopg is None:
        print("ERRO: pacote Python 'psycopg' não instalado. Instale antes de usar --apply.")
        sys.exit(1)

    print("Conectando ao banco de dados via DATABASE_URL...")

    inserted = 0
    updated = 0
    db_errors = []

    os.makedirs('data/output', exist_ok=True)

    try:
        with psycopg.connect(db_url) as conn:
            with conn.cursor() as cur:
                for item in parsed_data:
                    unidade = item['unidade']
                    financeiro = item['financeiro']

                    cur.execute(
                        "SELECT id FROM public.unidades_escolares WHERE designacao = %s",
                        (unidade['designacao'],),
                    )
                    already_exists = cur.fetchone() is not None

                    cur.execute(
                        """
                        INSERT INTO public.unidades_escolares (
                          designacao, nome, inep, cnpj, diretor, endereco,
                          agencia, conta_corrente, source_payload, ativo
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, true)
                        ON CONFLICT (designacao) DO UPDATE SET
                          nome = EXCLUDED.nome,
                          inep = EXCLUDED.inep,
                          cnpj = EXCLUDED.cnpj,
                          diretor = EXCLUDED.diretor,
                          endereco = EXCLUDED.endereco,
                          agencia = EXCLUDED.agencia,
                          conta_corrente = EXCLUDED.conta_corrente,
                          source_payload = EXCLUDED.source_payload,
                          ativo = true,
                          updated_at = now()
                        RETURNING id
                        """,
                        (
                            unidade['designacao'],
                            unidade['nome'],
                            unidade['inep'],
                            unidade['cnpj'],
                            unidade['diretor'],
                            unidade['endereco'],
                            unidade['agencia'],
                            unidade['conta_corrente'],
                            json.dumps(unidade['source_payload'], ensure_ascii=False),
                        ),
                    )
                    unidade_id = cur.fetchone()[0]

                    cur.execute(
                        """
                        INSERT INTO public.execucao_financeira (
                          unidade_id, exercicio, programa,
                          reprogramado_custeio, reprogramado_capital,
                          parcela_1_custeio, parcela_1_capital,
                          parcela_2_custeio, parcela_2_capital,
                          gasto
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
                        ON CONFLICT (unidade_id, exercicio, programa) DO UPDATE SET
                          reprogramado_custeio = EXCLUDED.reprogramado_custeio,
                          reprogramado_capital = EXCLUDED.reprogramado_capital,
                          parcela_1_custeio = EXCLUDED.parcela_1_custeio,
                          parcela_1_capital = EXCLUDED.parcela_1_capital,
                          parcela_2_custeio = EXCLUDED.parcela_2_custeio,
                          parcela_2_capital = EXCLUDED.parcela_2_capital,
                          updated_at = now()
                        """,
                        (
                            unidade_id,
                            financeiro['exercicio'],
                            financeiro['programa'],
                            financeiro['reprogramado_custeio'],
                            financeiro['reprogramado_capital'],
                            financeiro['parcela_1_custeio'],
                            financeiro['parcela_1_capital'],
                            financeiro['parcela_2_custeio'],
                            financeiro['parcela_2_capital'],
                        ),
                    )

                    if already_exists:
                        updated += 1
                    else:
                        inserted += 1

                cur.execute(
                    """
                    INSERT INTO public.import_logs (
                      source, filename, exercicio, programa, total_rows,
                      inserted_rows, updated_rows, skipped_rows, errors, status
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                    """,
                    (
                        'Script python --apply DATABASE_URL',
                        filepath,
                        exercicio,
                        programa,
                        total_valid + skipped,
                        inserted,
                        updated,
                        skipped,
                        '[]',
                        'sucesso',
                    ),
                )
    except Exception as e:
        db_errors.append({'erro': str(e)})
        all_errors = db_errors
        report = f"""# Relatório de Importação (Modo APPLY)
Data: {datetime.now().isoformat()}
Arquivo: {filepath}
Exercício: {exercicio} / {programa}
Inseridas no DB: 0
Atualizadas no DB: 0
Puladas vazias: {skipped}
Erros Totais (Validação + DB): {len(all_errors)}
Modo: FALHA TRANSACIONAL (nenhuma alteração confirmada)."""
        with open('data/output/import_report.md', 'w', encoding='utf-8') as f:
            f.write(report)
            f.write("\n\n## Erros\n")
            f.write(json.dumps(all_errors, ensure_ascii=False, indent=2))
        print(f"ERRO: falha transacional no modo --apply: {str(e)}")
        sys.exit(1)

    report = f"""# Relatório de Importação (Modo APPLY)
Data: {datetime.now().isoformat()}
Arquivo: {filepath}
Exercício: {exercicio} / {programa}
Inseridas no DB: {inserted}
Atualizadas no DB: {updated}
Puladas vazias: {skipped}
Erros Totais (Validação + DB): 0
Modo: GRAVADO NO BANCO DE DADOS."""
    with open('data/output/import_report.md', 'w', encoding='utf-8') as f:
        f.write(report)
        
    print("Modo Apply concluído. Relatório gerado.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Script de importação PDDE")
    parser.add_argument("filepath", help="Caminho do arquivo XLSX")
    parser.add_argument("--exercicio", type=int, default=2026)
    parser.add_argument("--programa", type=str, default='basico')
    parser.add_argument("--apply", action='store_true', help="Gravar no banco")
    parser.add_argument(
        "--approve-divergent-count",
        action='store_true',
        help="Permite --apply com total válido diferente de 163, após aprovação humana documentada",
    )
    
    args = parser.parse_args()
    process_file(args.filepath, args.exercicio, args.programa, args.apply, args.approve_divergent_count)
